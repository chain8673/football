import requests
from lxml import etree
import openpyxl

team_indexes = ['50001640', '50000087', '50001390', '50001146', '50000344', '50002040']
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36'}
filePath = 'E:\\ChenHao\\Python\\Project\\AsianTeams.xlsx'


def getTeamUrls(indexes):
    urls = []
    for index in indexes:
        url = 'https://www.dongqiudi.com/team/' + index + '.html'
        urls.append(url)
    return urls


def getDataFromUrl(url):
    response = requests.get(url=url, headers=headers)
    page_cont = response.content
    tree = etree.HTML(page_cont)
    players = tree.xpath('//*[@id="__layout"]/div/div[2]/div[2]/div/div[2]/div[2]/div[2]/p')
    player_tab = []
    for player in players:
        name = player.xpath('./a/span[3]/text()')[0].encode('iso-8859-1').decode()
        position = player.xpath('./a/span[1]/text()')[0].encode('iso-8859-1').decode()
        number = player.xpath('./a/span[2]/text()')[0].encode('iso-8859-1').decode()
        player = {'name': name, 'position': position, 'number': number}
        player_tab.append(player)
    return player_tab


def getTeamName(url):
    response = requests.get(url=url, headers=headers)
    page_cont = response.content
    tree = etree.HTML(page_cont)
    name = tree.xpath('//*[@id="__layout"]/div/div[2]/div[2]/div/div[1]/div/p[1]/text()')[0].encode(
        'iso-8859-1').decode()
    return name


def createExl(teamNames):
    wb = openpyxl.Workbook()
    for i in range(0, len(teamNames)):
        wb.create_sheet(teamNames[i], i)
        ws = wb[teamNames[i]]
        ws['A1'] = '位置'
        ws['B1'] = '姓名'
        ws['C1'] = '号码'
    wb.save(filePath)


def writePlayers(allPlayers, teamNames):
    wb = openpyxl.load_workbook(filePath)
    for i in range(len(allPlayers)):
        teamPlayers = allPlayers[i]
        ws = wb[teamNames[i]]
        for j in range(len(teamPlayers)):
            ws.append([teamPlayers[j]['position'], teamPlayers[j]['name'], teamPlayers[j]['number']])
    wb.save(filePath)


def main():
    team_urls = getTeamUrls(team_indexes)
    all_players = []
    team_names = []
    for team_url in team_urls:
        all_players.append(getDataFromUrl(team_url))
        team_names.append(getTeamName(team_url))
    for team_players in all_players:
        createExl(team_names)
    writePlayers(all_players, team_names)


if __name__ == '__main__':
    main()
